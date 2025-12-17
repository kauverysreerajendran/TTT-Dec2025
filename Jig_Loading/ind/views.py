# views.py
import json
from rest_framework import generics
from rest_framework.response import Response
from api.models import Patient
from . models import *
from api.serializers import PatientSerializer
from .serializers import *
from rest_framework import status
from rest_framework import viewsets
from rest_framework.decorators import action
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils.timezone import now, timedelta,datetime
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from django.views.decorators.csrf import csrf_exempt

class PatientDetailView(generics.RetrieveAPIView):
    queryset = Patient.objects.all()
    serializer_class = PatientSerializer

    def get(self, request, phone_number):
        try:
            patient = Patient.objects.get(patient_contact_number=phone_number)
            serializer = self.get_serializer(patient)
            return Response(serializer.data)
        except Patient.DoesNotExist:
            return Response({'error': 'Patient not found'}, status=404)

class VegetarianDietViewSet(viewsets.ModelViewSet):
    queryset = VegetarianDiet.objects.all()
    serializer_class = VegetarianDietSerializer

    def perform_create(self, serializer):
        serializer.save()

    def get_vegdiet_data(self, request, patient_id):
        today = timezone.now().date()  # Get today's date
        print(f"Today's date: {today}")  # Print today's date

        try:
            # Check if the patient exists and their diet preference
            patient = Patient.objects.get(patient_contact_number=patient_id)
            print(f"Patient Phone: {patient.patient_contact_number}, Diet: {patient.diet}")  # Print patient details

            if patient.diet not in ["Both", "Vegetarian"]:
                return Response({"error": "Patient diet preference does not include vegetarian"}, status=status.HTTP_400_BAD_REQUEST)

            # Filter the vegetarian diet by patient_id and today's date
            veg_diet = VegetarianDiet.objects.filter(patient_id=str(patient.patient_id), date=today)
            exists = veg_diet.exists()

            # Print debug logs
            print(f"Veg Diet Exists: {exists}, Patient Phone: {patient.patient_contact_number}, Date: {today}")

            return Response({
                "patient_id": patient.patient_id,
                "phone_number": patient.patient_contact_number,
                "diet": patient.diet,
                "date": str(today),
                "exists": exists
            }, status=status.HTTP_200_OK)

        except Patient.DoesNotExist:
            print(f"Error: Patient with phone {patient_id} not found!")
            return Response({"error": "Patient not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            print(f"Unexpected Error: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        
    def get_all_vegdiet_data(self, request, patient_id, date=None):
        # Use the provided date or default to today's date
        selected_date = date or timezone.now().date()
        try:
            # Ensure the date is in the correct format (YYYY-MM-DD)
            selected_date = timezone.datetime.strptime(selected_date, "%Y-%m-%d").date()
            
            # Filter the sleep rituals by patient_id and the selected date
            all_veg_diet = VegetarianDiet.objects.filter(patient_id=patient_id, date=selected_date)
            
            if all_veg_diet.exists():
                return Response({"exists": True}, status=status.HTTP_200_OK)
            else:
                return Response({"exists": False}, status=status.HTTP_200_OK)
        except ValueError:
            return Response({"error": "Invalid date format. Please use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        

class NonVegetarianDietViewSet(viewsets.ModelViewSet):
    queryset = NonVegetarianDiet.objects.all()
    serializer_class = NonVegetarianDietSerializer

    def perform_create(self, serializer):
        serializer.save()
        
    def get_nonvegdiet_data(self, request, patient_id):
        today = timezone.now().date()  # Get today's date
        print(f"Today's date: {today}")  # Debugging log

        try:
            # Get the patient details
            patient = Patient.objects.get(patient_contact_number=patient_id)
            print(f"Patient ID: {patient.patient_id}, Diet: {patient.diet}, Phone: {patient.patient_contact_number}")

            if patient.diet not in ["Both", "Non-Vegetarian"]:
                return Response({"error": "Patient diet preference does not include nonvegetarian"}, status=status.HTTP_400_BAD_REQUEST)

            # Ensure correct patient_id format before querying NonVegetarianDiet
            nonveg_diet = NonVegetarianDiet.objects.filter(patient_id=str(patient.patient_id), date=today)
            exists = nonveg_diet.exists()

            print(f"Non-Veg Diet Exists: {exists}, Patient Phone: {patient.patient_contact_number}, Date: {today}")

            return Response({
                "patient_id": patient.patient_id,
                "phone_number": patient.patient_contact_number,
                "diet": patient.diet,
                "date": str(today),
                "exists": exists
            }, status=status.HTTP_200_OK)

        except Patient.DoesNotExist:
            print(f"Error: Patient with phone {patient_id} not found!")
            return Response({"error": "Patient not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            print(f"Unexpected Error: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        
    def get_all_nonvegdiet_data(self, request, patient_id, date=None):
        # Use the provided date or default to today's date
        selected_date = date or timezone.now().date()
        try:
            # Ensure the date is in the correct format (YYYY-MM-DD)
            selected_date = timezone.datetime.strptime(selected_date, "%Y-%m-%d").date()
            
            # Filter the sleep rituals by patient_id and the selected date
            all_nonveg_diet = NonVegetarianDiet.objects.filter(patient_id=patient_id, date=selected_date)
            
            if all_nonveg_diet.exists():
                return Response({"exists": True}, status=status.HTTP_200_OK)
            else:
                return Response({"exists": False}, status=status.HTTP_200_OK)
        except ValueError:
            return Response({"error": "Invalid date format. Please use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
    

class SleepRitualViewSet(viewsets.ModelViewSet):
    queryset = SleepRitual.objects.all()
    serializer_class = SleepRitualSerializer

    def perform_create(self, serializer):
        serializer.save()

    def get_sleep_data(self, request, patient_id):
        today = timezone.now().date()  # Get today's date
        try:
            # Filter the sleep rituals by patient_id and today's date
            sleep_rituals = SleepRitual.objects.filter(patient_id=patient_id, date=today)
            if sleep_rituals.exists():
                return Response({"exists": True}, status=status.HTTP_200_OK)
            else:
                return Response({"exists": False}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
    def get_all_sleep_data(self, request, patient_id, date=None):
        # Use the provided date or default to today's date
        selected_date = date or timezone.now().date()
        try:
            # Ensure the date is in the correct format (YYYY-MM-DD)
            selected_date = timezone.datetime.strptime(selected_date, "%Y-%m-%d").date()
            
            # Filter the sleep rituals by patient_id and the selected date
            sleep_rituals = SleepRitual.objects.filter(patient_id=patient_id, date=selected_date)
            
            # Check if sleep rituals exist for the specified date
            if sleep_rituals.exists():
                return Response({"exists": True}, status=status.HTTP_200_OK)
            else:
                return Response({"exists": False}, status=status.HTTP_200_OK)
        except ValueError:
            return Response({"error": "Invalid date format. Please use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
class WaterIntakeViewSet(viewsets.ModelViewSet):
    queryset = WaterIntake.objects.all()
    serializer_class = WaterIntakeSerializer

    def perform_create(self, serializer):
        serializer.save()

    def get_water_data(self, request, patient_id):
        today = timezone.now().date()  # Get today's date
        try:
            # Filter the sleep rituals by patient_id and today's date
            water_data = WaterIntake.objects.filter(patient_id=patient_id, date=today)
            if water_data.exists():
                return Response({"exists": True}, status=status.HTTP_200_OK)
            else:
                return Response({"exists": False}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
    def get_all_water_data(self, request, patient_id, date=None):
        # Use the provided date or default to today's date
        selected_date = date or timezone.now().date()
        try:
            # Ensure the date is in the correct format (YYYY-MM-DD)
            selected_date = timezone.datetime.strptime(selected_date, "%Y-%m-%d").date()
            
            # Filter the sleep rituals by patient_id and the selected date
            water_datas = WaterIntake.objects.filter(patient_id=patient_id, date=selected_date)
            
            # Check if sleep rituals exist for the specified date
            if water_datas.exists():
                return Response({"exists": True}, status=status.HTTP_200_OK)
            else:
                return Response({"exists": False}, status=status.HTTP_200_OK)
        except ValueError:
            return Response({"error": "Invalid date format. Please use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
class DailyExerciseViewSet(viewsets.ModelViewSet):
    queryset = DailyExercise.objects.all()
    serializer_class = DailyExerciseSerializer

    def perform_create(self, serializer):
        serializer.save()

    def get_daily_exercise_data(self, request, patient_id):
        today = timezone.now().date()  # Get today's date
        try:
            # Filter the sleep rituals by patient_id and today's date
            daily_exercise_data = DailyExercise.objects.filter(patient_id=patient_id, date=today)
            if daily_exercise_data.exists():
                return Response({"exists": True}, status=status.HTTP_200_OK)
            else:
                return Response({"exists": False}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
    def get_all_daliyexercise_data(self, request, patient_id, date=None):
        # Use the provided date or default to today's date
        selected_date = date or timezone.now().date()
        try:
            # Ensure the date is in the correct format (YYYY-MM-DD)
            selected_date = timezone.datetime.strptime(selected_date, "%Y-%m-%d").date()
            
            # Filter the sleep rituals by patient_id and the selected date
            daliyexercise_datas = DailyExercise.objects.filter(patient_id=patient_id, date=selected_date)
            
            # Check if sleep rituals exist for the specified date
            if daliyexercise_datas.exists():
                return Response({"exists": True}, status=status.HTTP_200_OK)
            else:
                return Response({"exists": False}, status=status.HTTP_200_OK)
        except ValueError:
            return Response({"error": "Invalid date format. Please use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
class WalkingActivityViewSet(viewsets.ModelViewSet):
    queryset = WalkingActivity.objects.all()
    serializer_class = WalkingActivitySerializer

    def perform_create(self, serializer):
        # Save the walking activity with the user-provided data
        serializer.save()

    def get_walking_data(self, request, patient_id):
        today = timezone.now().date()  # Get today's date
        try:
            # Filter the sleep rituals by patient_id and today's date
            walking_data = WalkingActivity.objects.filter(patient_id=patient_id, date=today)
            if walking_data.exists():
                return Response({"exists": True}, status=status.HTTP_200_OK)
            else:
                return Response({"exists": False}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
    def get_all_walking_data(self, request, patient_id, date=None):
        # Use the provided date or default to today's date
        selected_date = date or timezone.now().date()
        try:
            # Ensure the date is in the correct format (YYYY-MM-DD)
            selected_date = timezone.datetime.strptime(selected_date, "%Y-%m-%d").date()
            
            # Filter the sleep rituals by patient_id and the selected date
            walking_datas = WalkingActivity.objects.filter(patient_id=patient_id, date=selected_date)
            
            # Check if sleep rituals exist for the specified date
            if walking_datas.exists():
                return Response({"exists": True}, status=status.HTTP_200_OK)
            else:
                return Response({"exists": False}, status=status.HTTP_200_OK)
        except ValueError:
            return Response({"error": "Invalid date format. Please use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
    from .serializers import WalkingActivitySerializer

    @action(detail=False, methods=["get"])
    def get_current_week_data(self, request, patient_id, date):
        try:
            # Parse the date
            selected_date = timezone.datetime.strptime(date, "%Y-%m-%d").date()

            # Calculate the start and end of the week for the given date
            start_of_week = selected_date - timedelta(days=selected_date.weekday())  # Monday
            end_of_week = start_of_week + timedelta(days=6)  # Sunday

            # Query for walking activities within the week for the given patient
            walking_activities = WalkingActivity.objects.filter(
                patient_id=patient_id,
                date__range=[start_of_week, end_of_week]
            )

            # If no walking activities exist for the week, return 0 km data
            if not walking_activities.exists():
                walking_activities = [{
                    "patient_id": patient_id,
                    "date": selected_date,
                    "distance_km": 0.0,  # Set distance to 0 km
                    "duration_hours": 0,
                    "duration_minutes": 0,
                    "difficulty_text": "No data for this week"
                }]

            # Serialize the walking activity data
            serialized_data = WalkingActivitySerializer(walking_activities, many=True)

            # Return the response with walking activities
            return Response(
                {
                    "start_of_week": start_of_week,
                    "end_of_week": end_of_week,
                    "walking_activities": serialized_data.data,
                },
                status=status.HTTP_200_OK,
            )

        except ValueError:
            return Response({"error": "Invalid date format. Please use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class YogaActivityViewSet(viewsets.ModelViewSet):
    queryset = YogaActivity.objects.all()
    serializer_class = YogaActivitySerializer

    def perform_create(self, serializer):
        # Save the walking activity with the user-provided data
        serializer.save()

    def get_yoga_data(self, request, patient_id):
        today = timezone.now().date()  # Get today's date
        try:
            # Filter the sleep rituals by patient_id and today's date
            yoga_data = YogaActivity.objects.filter(patient_id=patient_id, date=today)
            if yoga_data.exists():
                return Response({"exists": True}, status=status.HTTP_200_OK)
            else:
                return Response({"exists": False}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
    def get_all_yoga_data(self, request, patient_id, date=None):
        # Use the provided date or default to today's date
        selected_date = date or timezone.now().date()
        try:
            # Ensure the date is in the correct format (YYYY-MM-DD)
            selected_date = timezone.datetime.strptime(selected_date, "%Y-%m-%d").date()
            
            # Filter the sleep rituals by patient_id and the selected date
            yoga_datas = YogaActivity.objects.filter(patient_id=patient_id, date=selected_date)
            
            # Check if sleep rituals exist for the specified date
            if yoga_datas.exists():
                return Response({"exists": True}, status=status.HTTP_200_OK)
            else:
                return Response({"exists": False}, status=status.HTTP_200_OK)
        except ValueError:
            return Response({"error": "Invalid date format. Please use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
class LifeStyleTrackerViewSet(viewsets.ModelViewSet):
    queryset = LifeStyleTracker.objects.all()
    serializer_class = LifeStyleTrackerSerializer

    def perform_create(self, serializer):
        # Save the walking activity with the user-provided data
        serializer.save()

    def get_lifestyle_data(self, request, patient_id):
        today = timezone.now().date()  # Get today's date
        try:
            # Filter the sleep rituals by patient_id and today's date
            liftstyle_data = LifeStyleTracker.objects.filter(patient_id=patient_id, date=today)
            if liftstyle_data.exists():
                return Response({"exists": True}, status=status.HTTP_200_OK)
            else:
                return Response({"exists": False}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
    def get_all_lifestyle_data(self, request, patient_id, date=None):
        # Use the provided date or default to today's date
        selected_date = date or timezone.now().date()
        try:
            # Ensure the date is in the correct format (YYYY-MM-DD)
            selected_date = timezone.datetime.strptime(selected_date, "%Y-%m-%d").date()
            
            # Filter the sleep rituals by patient_id and the selected date
            all_lifestyle_datas = LifeStyleTracker.objects.filter(patient_id=patient_id, date=selected_date)
            
            # Check if sleep rituals exist for the specified date
            if all_lifestyle_datas.exists():
                return Response({"exists": True}, status=status.HTTP_200_OK)
            else:
                return Response({"exists": False}, status=status.HTTP_200_OK)
        except ValueError:
            return Response({"error": "Invalid date format. Please use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        

class SubmissionCountViewSet(viewsets.ViewSet):
    def get_notification_count(self, request, patient_id):
        today = timezone.now().date()  # Get today's date
        try:
            patient = Patient.objects.get(patient_id=patient_id)

            # Count entries for each model
            veg_diet_count = VegetarianDiet.objects.filter(patient_id=patient_id, date=today).count()
            nonveg_diet_count = NonVegetarianDiet.objects.filter(patient_id=patient_id, date=today).count()
            sleep_ritual_count = SleepRitual.objects.filter(patient_id=patient_id, date=today).count()
            water_intake_count = WaterIntake.objects.filter(patient_id=patient_id, date=today).count()
            daily_exercise_count = DailyExercise.objects.filter(patient_id=patient_id, date=today).count()
            walking_activity_count = WalkingActivity.objects.filter(patient_id=patient_id, date=today).count()
            yoga_activity_count = YogaActivity.objects.filter(patient_id=patient_id, date=today).count()
            medicine_count = MedicationPatient.objects.filter(patient_id=patient_id, date=today).count()
            lifestyle_count = LifeStyleTracker.objects.filter(patient_id=patient_id, date=today).count()

            # Exclude diet count based on patient diet type
            diet_counts = {}
            if patient.diet != "Vegetarian":
                diet_counts["nonveg_diet_count"] = 1 if nonveg_diet_count == 0 else 0
            if patient.diet != "Non-Vegetarian":
                diet_counts["veg_diet_count"] = 1 if veg_diet_count == 0 else 0

            # Individual not submitted count
            individual_counts = {
                "veg_diet_count": 1 if veg_diet_count == 0 else 0,
                "nonveg_diet_count": 1 if nonveg_diet_count == 0 else 0,
                "sleep_ritual_count": 1 if sleep_ritual_count == 0 else 0,
                "water_intake_count": 1 if water_intake_count == 0 else 0,
                "daily_exercise_count": 1 if daily_exercise_count == 0 else 0,
                "walking_activity_count": 1 if walking_activity_count == 0 else 0,
                "yoga_activity_count": 1 if yoga_activity_count == 0 else 0,
                "lifestyle_count": 1 if lifestyle_count == 0 else 0,
                "medicine_count": 1 if medicine_count == 0 else 0
            }

            # Apply diet filter and calculate the total count
            total_not_submitted = sum(diet_counts.values()) + sum(
                v for k, v in individual_counts.items() if k not in ["veg_diet_count", "nonveg_diet_count"]
            )

            # Print Debugging Information
            print(f"Patient ID: {patient_id}")
            print(f"Diet: {patient.diet}")
            print(f"Individual Counts: {individual_counts}")
            print(f"Filtered Diet Counts: {diet_counts}")
            print(f"Overall Not Submitted Count: {total_not_submitted}")

            return Response({
                "patient_id": patient_id,
                "diet": patient.diet,
                "individual_counts": individual_counts,
                "diet_counts": diet_counts,
                "notification_count": total_not_submitted
            }, status=status.HTTP_200_OK)

        except Patient.DoesNotExist:
            return Response({"error": "Patient not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class UserFeedbackViewSet(viewsets.ModelViewSet):
    queryset = UserFeedback.objects.all()
    serializer_class = UserFeedbackSerializer

    def perform_create(self, serializer):
        # Save the walking activity with the user-provided data
        serializer.save()
    

from io import BytesIO


def export_patient_data_to_excel(request):
    # Create a new Excel workbook and a worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Patient Data"

    # Define headers
    headers = [
        "Patient ID", "Date", "Sleep Data", "Vegetarian Diet", "Non-Vegetarian Diet",
        "Water Intake", "Daily Exercise", "Walking Activity", "Yoga Activity","Medicine","Lifestyle"
    ]
    sheet.append(headers)

    # Fetch data from each model and append to sheet
    patient_data = []

    veg_diets = VegetarianDiet.objects.all()
    non_veg_diets = NonVegetarianDiet.objects.all()
    sleep_rituals = SleepRitual.objects.all()
    water_intakes = WaterIntake.objects.all()
    exercises = DailyExercise.objects.all()
    walks = WalkingActivity.objects.all()
    yoga_activities = YogaActivity.objects.all()
    medicine_activities = MedicationPatient.objects.all()
    lifestyle_activities = LifeStyleTracker.objects.all()

    for veg_diet in veg_diets:
        row_data = [
            veg_diet.patient_id, veg_diet.date,
            get_data(sleep_rituals, veg_diet.patient_id, veg_diet.date, "sleep"),
            get_data(veg_diets, veg_diet.patient_id, veg_diet.date, "veg_diet"),
            get_data(non_veg_diets, veg_diet.patient_id, veg_diet.date, "non_veg_diet"),
            get_data(water_intakes, veg_diet.patient_id, veg_diet.date, "water"),
            get_data(exercises, veg_diet.patient_id, veg_diet.date, "exercise"),
            get_data(walks, veg_diet.patient_id, veg_diet.date, "walk"),
            get_data(yoga_activities, veg_diet.patient_id, veg_diet.date, "yoga"),
            get_data(medicine_activities, veg_diet.patient_id, veg_diet.date, "medicine"),
            get_data(lifestyle_activities, veg_diet.patient_id, veg_diet.date, "lifestyle")

        ]
        print("Row data:", row_data)  # Log row data for debugging

        sheet.append(row_data)

    # Set column widths
    for col in range(1, sheet.max_column + 1):
        col_letter = get_column_letter(col)
        sheet.column_dimensions[col_letter].width = 20

    # Save the workbook to a BytesIO object
    output = BytesIO()
    workbook.save(output)
    output.seek(0)  # Move the cursor to the start of the stream

    # Prepare the response
    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="patient_data.xlsx"'

    return response

def get_data(queryset, patient_id, date, data_type):
    try:
        entries = queryset.filter(patient_id=patient_id, date=date)
        if not entries.exists():
            return "No data available"

        # If multiple entries are found, aggregate or concatenate the data
        if data_type == "sleep":
            return "; ".join(
                f"Warm Milk: {entry.warm_milk}, Sleep Breaks: {entry.sleep_breaks}"
                for entry in entries
            )
        elif data_type == "veg_diet":
            return "; ".join(
                f"Sprouts: {entry.sprouts_quantity}, Guava: {entry.guava_quantity}"
                for entry in entries
            )
        elif data_type == "non_veg_diet":
            return "; ".join(
                f"Eggs: {entry.eggs_quantity}, Fish: {entry.fish_quantity}"
                for entry in entries
            )
        elif data_type == "water":
            return "; ".join(
                f"1000ml: {entry.ml_1000}, 2000ml: {entry.ml_2000}" for entry in entries
            )
        elif data_type == "exercise":
            return "; ".join(
                f"Warm-Up: {entry.warm_up}, Chores: {entry.chores}" for entry in entries
            )
        elif data_type == "walk":
            return "; ".join(
                f"Distance: {entry.distance_km} km, Duration: {entry.duration_hours}h {entry.duration_minutes}m"
                for entry in entries
            )
        elif data_type == "yoga":
            return "; ".join(
                f"Mindful: {entry.mindful_yoga}, Duration: {entry.duration_hours}h {entry.duration_minutes}m"
                for entry in entries
            )
        elif data_type == "lifestyle":
            return "; ".join(
                f"Smoking: {entry.smoking}, No of Cigarettes: {entry.no_of_cigraretes}, "
                f"Alcoholic: {entry.alcoholic}, Pastries: {entry.pastries}, Sweets: {entry.sweets}, "
                f"Salted Items: {entry.salted_items}, Preserved Foods: {entry.preserved_foods}, "
                f"Pickles: {entry.pickles}, Dry Fish: {entry.dry_fish}, Fried Food: {entry.fried_food}, "
                f"Processed Meat: {entry.processed_meat}, Others: {entry.others_track}"
                for entry in entries
            )
        elif data_type == "medication":
            return "; ".join(
                f"Medication Name: {entry.medication_name}, Route: {entry.route}, Dosage Amount: {entry.dosage_amount} {entry.dosage_type}, "
                f"Drug Take: {entry.drug_take}, Consume Instructions: {entry.consume}, Drug Action: {entry.drug_action}"
                for entry in entries
            )
        else:
            return "Invalid data type"
    except Exception as e:
        return f"Error: {str(e)}"




class MedicationPatientViewSet(viewsets.ModelViewSet):
    queryset = MedicationPatient.objects.all()
    serializer_class = MedicationPatientSerializer

    def create(self, request, *args, **kwargs):
        # Here we expect a list of medications
        serializer = self.get_serializer(data=request.data, many=True)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    def get_medicine_data(self, request, patient_id):
        today = timezone.now().date()  # Get today's date
        try:
            # Filter the sleep rituals by patient_id and today's date
            medicine_data = MedicationPatient.objects.filter(patient_id=patient_id, date=today)
            if medicine_data.exists():
                return Response({"exists": True}, status=status.HTTP_200_OK)
            else:
                return Response({"exists": False}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
    def get_all_medicine_data(self, request, patient_id, date=None):
        # Use the provided date or default to today's date
        selected_date = date or timezone.now().date()
        try:
            # Ensure the date is in the correct format (YYYY-MM-DD)
            selected_date = timezone.datetime.strptime(selected_date, "%Y-%m-%d").date()
            
            # Filter the sleep rituals by patient_id and the selected date
            medicine_datas = MedicationPatient.objects.filter(patient_id=patient_id, date=selected_date)
            
            # Check if sleep rituals exist for the specified date
            if medicine_datas.exists():
                return Response({"exists": True}, status=status.HTTP_200_OK)
            else:
                return Response({"exists": False}, status=status.HTTP_200_OK)
        except ValueError:
            return Response({"error": "Invalid date format. Please use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
    def get_week_medicine_data(self, request, patient_id, date):
        """
        Fetch medication data for a week (Monday to Sunday) and calculate days taken/missed.
        """
        try:
            # Parse the provided date and find the week's start (Monday)
            provided_date = datetime.strptime(date, "%Y-%m-%d").date()
            week_start = provided_date - timedelta(days=provided_date.weekday())
            week_end = week_start + timedelta(days=6)

            # Query medications for the week
            week_data = MedicationPatient.objects.filter(
                patient_id=patient_id,
                date__range=[week_start, week_end]
            )

            # Initialize the weekly status with default "Yet to Take"
            week_status = {week_start + timedelta(days=i): "Yet to Take" for i in range(7)}
            day_names = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']

            # Determine the current date and set the status
            current_date = datetime.now().date()

            # Update the status based on the medication data for the week
            for record in week_data:
                if record.taken:
                    week_status[record.date] = "Taken"
                else:
                    week_status[record.date] = "Missed"

            # Update the status of future days (after today) to "Yet to Take"
            for i in range(7):
                current_day = week_start + timedelta(days=i)
                if current_day > current_date:
                    week_status[current_day] = "Yet to Take"
                elif current_day < current_date:
                    # For past days, check if the status is still "Yet to Take" and set to "Missed" if not taken
                    if week_status[current_day] == "Yet to Take":
                        week_status[current_day] = "Missed"

            # Prepare the response data
            response_data = {
                "week_start": week_start,
                "week_end": week_end,
                "data": [{"day": day_names[i], "status": week_status[week_start + timedelta(days=i)]} for i in range(7)],
                "summary": {
                    "taken": list(week_status.values()).count("Taken"),
                    "missed": list(week_status.values()).count("Missed"),
                    "yet_to_take": list(week_status.values()).count("Yet to Take")
                }
            }

            return Response(response_data, status=status.HTTP_200_OK)

        except ValueError:
            return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


from django.http import JsonResponse
import json
from .firebase import send_push_notification

@csrf_exempt
def notify_user(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        token = data.get('token')
        title = data.get('title')
        message = data.get('message')
        response = send_push_notification(token, title, message)
        return JsonResponse({'status': 'success', 'response': response})
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)


@csrf_exempt
def save_token(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        token = data.get('token')
        patient_id = data.get('patient_id')
        group_name = 'Patient'  # Set group_name to 'Patient'
        
        if token and group_name and patient_id:
            FCMToken.objects.update_or_create(token=token, defaults={'group_name': group_name, 'patient_id': patient_id})
            return JsonResponse({'status': 'success', 'message': 'Token, group name, and patient saved successfully'})
        return JsonResponse({'status': 'error', 'message': 'Token, group name, and patient ID are required'}, status=400)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)

